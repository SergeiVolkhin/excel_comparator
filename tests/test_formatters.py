"""Characterization tests for ExcelOutputFormatter and HTMLOutputFormatter."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.comparators.advanced_comparator import AdvancedComparator
from src.comparators.basic_comparator import BasicComparator
from src.core.interfaces import ComparisonResult
from src.formatters.excel_formatter import (
    COLUMN_WIDTH_SAMPLE_ROWS,
    EXCEL_LARGE_ROW_THRESHOLD,
    ExcelOutputFormatter,
)
from src.formatters.html_formatter import HTMLOutputFormatter


@pytest.fixture
def simple_result() -> ComparisonResult:
    df1 = pd.DataFrame({"id": [1, 2, 3], "name": ["a", "b", "c"]})
    df2 = pd.DataFrame({"id": [1, 2, 3], "name": ["a", "B", "c"]})
    return BasicComparator().compare(df1, df2)


class TestExcelFormatter:
    def test_format_writes_xlsx(self, simple_result: ComparisonResult, tmp_path: Path) -> None:
        out = tmp_path / "r.xlsx"
        ExcelOutputFormatter().format(simple_result, out, file1_name="A", file2_name="B")
        assert out.exists() and out.stat().st_size > 0

    def test_format_produces_expected_sheets(
        self, simple_result: ComparisonResult, tmp_path: Path
    ) -> None:
        out = tmp_path / "r.xlsx"
        ExcelOutputFormatter().format(simple_result, out, file1_name="FileA", file2_name="FileB")
        wb = load_workbook(out, read_only=True)
        try:
            # Two data sheets (file stems, possibly truncated) + a summary sheet
            assert len(wb.sheetnames) == 3
        finally:
            wb.close()

    def test_supported_formats(self) -> None:
        assert ExcelOutputFormatter().get_supported_formats() == [".xlsx"]

    def test_different_row_counts_writes_na_as_empty(self, tmp_path: Path) -> None:
        # Advanced alignment pads the shorter frame with pd.NA, which openpyxl
        # cannot write ("Cannot convert <NA> to Excel"). The formatter must
        # turn those cells into empty (None) cells instead of crashing.
        df1 = pd.DataFrame({"id": [1, 2, 3], "name": ["a", "b", "c"]})
        df2 = pd.DataFrame({"id": [1, 2], "name": ["a", "b"]})
        result = AdvancedComparator().compare(df1, df2)

        out = tmp_path / "padded.xlsx"
        ExcelOutputFormatter().format(result, out, file1_name="A", file2_name="B")

        wb = load_workbook(out, read_only=True)
        try:
            sheet_b = wb["B"]  # the shorter file2 sheet, padded with NA
            # Header + 3 rows (padded to the longer frame's length).
            rows = list(sheet_b.iter_rows(values_only=True))
        finally:
            wb.close()
        # The 3rd data row was padding: its source cells (id, name) must be
        # empty (None), not <NA>. A trailing "Различия" column describes the diff.
        assert rows[3][0] is None
        assert rows[3][1] is None

    def test_column_width_samples_only_first_rows(self, tmp_path: Path) -> None:
        # A very long value placed BEYOND the sampled window must not widen
        # the column — proves _auto_adjust_columns no longer scans every row.
        n = COLUMN_WIDTH_SAMPLE_ROWS + 10
        names = ["a"] * n
        names[-1] = "X" * 100  # outside header + first COLUMN_WIDTH_SAMPLE_ROWS
        df = pd.DataFrame({"id": range(n), "name": names})
        result = BasicComparator().compare(df, df.copy())

        out = tmp_path / "wide.xlsx"
        ExcelOutputFormatter().format(result, out, file1_name="A", file2_name="B")

        wb = load_workbook(out)
        try:
            width_b = wb["A"].column_dimensions["B"].width  # the "name" column
        finally:
            wb.close()
        # Sampled width reflects "name"/"a" (~6). A full scan would cap at 50.
        assert width_b is not None
        assert width_b < 20

    def test_large_row_count_logs_html_suggestion(
        self, simple_result: ComparisonResult, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        simple_result.metadata["shape"] = (EXCEL_LARGE_ROW_THRESHOLD + 1, 2)
        out = tmp_path / "big.xlsx"
        with caplog.at_level(logging.WARNING, logger="ExcelOutputFormatter"):
            ExcelOutputFormatter().format(simple_result, out, file1_name="A", file2_name="B")
        assert any(".html" in r.message for r in caplog.records)


class TestHTMLFormatter:
    def test_format_writes_html(self, simple_result: ComparisonResult, tmp_path: Path) -> None:
        out = tmp_path / "r.html"
        HTMLOutputFormatter().format(simple_result, out, file1_name="A", file2_name="B")
        assert out.exists()
        content = out.read_text(encoding="utf-8")
        assert "<html" in content.lower() or "<!doctype" in content.lower()

    def test_supported_formats(self) -> None:
        fmts = HTMLOutputFormatter().get_supported_formats()
        assert ".html" in fmts or ".htm" in fmts

    def test_paginated_report(self, tmp_path: Path) -> None:
        # Force pagination by lowering the page_size to a tiny value.
        formatter = HTMLOutputFormatter()
        formatter.page_size = 10
        df1 = pd.DataFrame({"a": list(range(25))})
        df2 = pd.DataFrame({"a": list(range(25))})
        result = BasicComparator().compare(df1, df2)
        out = tmp_path / "r.html"
        formatter.format(result, out, file1_name="A", file2_name="B")
        # Pagination creates sibling page_*.html files
        pages = list(tmp_path.glob("r_page_*.html"))
        assert len(pages) >= 2
        assert out.exists()
