"""Shared pytest fixtures: generate test xlsx files on the fly.

All xlsx artefacts live in tmp_path and are not committed. Small sample
DataFrames are produced by named factories. A helper `xlsx_to_dict` reads
an xlsx back into a structure suitable for snapshot comparison (no zip
timestamp noise).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------


def _write_xlsx(path: Path, data: dict[str, pd.DataFrame]) -> Path:
    """Write a {sheet_name: DataFrame} mapping to an xlsx file."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    return path


def _write_raw_xlsx(path: Path, sheets: dict[str, list[list[Any]]]) -> Path:
    """Write raw rows (including formulas) via openpyxl."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Snapshot helper
# ---------------------------------------------------------------------------


def xlsx_to_dict(path: Path) -> dict[str, list[list[Any]]]:
    """Read xlsx into a stable dict for snapshot comparison."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        return {
            sheet: [list(row) for row in wb[sheet].iter_rows(values_only=True)]
            for sheet in wb.sheetnames
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Base data factories
# ---------------------------------------------------------------------------


def _df_base() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "id": [1, 2, 3, 4],
            "name": ["Alice", "Bob", "Charlie", "Diana"],
            "score": [10.5, 20.0, 30.25, 40.75],
        }
    )


def _df_value_diff() -> pd.DataFrame:
    df = _df_base()
    df.loc[1, "name"] = "Robert"
    df.loc[2, "score"] = 99.99
    return df


def _df_empty_cells() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "id": [1, 2, 3],
            "name": ["Alice", None, "Charlie"],
            "score": [10.5, 20.0, None],
        }
    )


def _df_mixed_types() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "id": [1, 2, 3],
            "value": [100, "100", 100.0],
            "flag": [True, False, True],
        }
    )


# ---------------------------------------------------------------------------
# Pytest fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def xlsx_dir(tmp_path: Path) -> Path:
    d = tmp_path / "xlsx"
    d.mkdir()
    return d


@pytest.fixture
def identical_xlsx_pair(xlsx_dir: Path) -> tuple[Path, Path]:
    df = _df_base()
    a = _write_xlsx(xlsx_dir / "a.xlsx", {"Sheet1": df})
    b = _write_xlsx(xlsx_dir / "b.xlsx", {"Sheet1": df})
    return a, b


@pytest.fixture
def value_diff_xlsx_pair(xlsx_dir: Path) -> tuple[Path, Path]:
    a = _write_xlsx(xlsx_dir / "a.xlsx", {"Sheet1": _df_base()})
    b = _write_xlsx(xlsx_dir / "b.xlsx", {"Sheet1": _df_value_diff()})
    return a, b


@pytest.fixture
def empty_cells_xlsx_pair(xlsx_dir: Path) -> tuple[Path, Path]:
    a = _write_xlsx(xlsx_dir / "a.xlsx", {"Sheet1": _df_base().head(3)})
    b = _write_xlsx(xlsx_dir / "b.xlsx", {"Sheet1": _df_empty_cells()})
    return a, b


@pytest.fixture
def mixed_types_xlsx_pair(xlsx_dir: Path) -> tuple[Path, Path]:
    a = _write_xlsx(xlsx_dir / "a.xlsx", {"Sheet1": _df_mixed_types()})
    df_b = _df_mixed_types().copy()
    df_b.loc[0, "flag"] = False
    b = _write_xlsx(xlsx_dir / "b.xlsx", {"Sheet1": df_b})
    return a, b


@pytest.fixture
def multi_sheet_xlsx_pair(xlsx_dir: Path) -> tuple[Path, Path]:
    data_a = {"Main": _df_base(), "Extra": _df_base().head(2)}
    data_b = {"Main": _df_value_diff(), "Extra": _df_base().head(2)}
    a = _write_xlsx(xlsx_dir / "a.xlsx", data_a)
    b = _write_xlsx(xlsx_dir / "b.xlsx", data_b)
    return a, b


@pytest.fixture
def formulas_xlsx_pair(xlsx_dir: Path) -> tuple[Path, Path]:
    rows_a = [["id", "x", "y", "sum"], [1, 2, 3, "=B2+C2"], [2, 5, 7, "=B3+C3"]]
    rows_b = [["id", "x", "y", "sum"], [1, 2, 3, "=B2+C2"], [2, 5, 8, "=B3+C3"]]
    a = _write_raw_xlsx(xlsx_dir / "a.xlsx", {"Sheet1": rows_a})
    b = _write_raw_xlsx(xlsx_dir / "b.xlsx", {"Sheet1": rows_b})
    return a, b


@pytest.fixture
def large_xlsx_pair(xlsx_dir: Path) -> tuple[Path, Path]:
    """50k rows smoke fixture — marked via @pytest.mark.slow at call site."""
    n = 50_000
    df_a = pd.DataFrame({"id": range(n), "v": range(n)})
    df_b = df_a.copy()
    df_b.loc[n // 2, "v"] = -1  # single diff deep in the file
    a = _write_xlsx(xlsx_dir / "a.xlsx", {"Sheet1": df_a})
    b = _write_xlsx(xlsx_dir / "b.xlsx", {"Sheet1": df_b})
    return a, b
