"""
Форматтер для вывода результатов в Excel
"""

import logging
from pathlib import Path
from typing import Any, ClassVar

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..analyzers.list_analyzer import BasicDifferenceAnalyzer, ListDifferenceAnalyzer
from ..core.exceptions import ApplicationError
from ..core.interfaces import ComparisonResult, IDifferenceAnalyzer, IOutputFormatter

#: Excel spec caps sheet names at 31 chars; we stay one under to avoid
#: silent truncation on append when duplicate-suffix resolution kicks in.
SHEET_NAME_MAX_LEN: int = 30

#: Default hex color used for highlighting differing cells.
DEFAULT_HIGHLIGHT_COLOR: str = "FFFF00"

#: Header row background color.
HEADER_FILL_COLOR: str = "E6E6FA"

#: Column-width auto-fit samples only the first N data rows instead of the
#: whole sheet. Width is a visual estimate, so sampling keeps the cost roughly
#: constant rather than O(rows × cols) on large tables.
COLUMN_WIDTH_SAMPLE_ROWS: int = 200

#: Above this row count an Excel report is slow to build; we log a warning
#: suggesting the HTML formatter (which paginates) for large tables.
EXCEL_LARGE_ROW_THRESHOLD: int = 20_000


class ExcelOutputFormatter(IOutputFormatter):
    """Форматтер для вывода результатов сравнения в Excel"""

    SUPPORTED_FORMATS: ClassVar[list[str]] = [".xlsx"]

    def __init__(self, config: Any = None) -> None:
        self.logger = logging.getLogger(self.__class__.__name__)
        self.config = config

        # Настройки стилей
        highlight_color = (
            getattr(config.comparison, "highlight_color", DEFAULT_HIGHLIGHT_COLOR)
            if config
            else DEFAULT_HIGHLIGHT_COLOR
        )
        self.highlight_fill = PatternFill(
            start_color=highlight_color,
            end_color=highlight_color,
            fill_type="solid",
        )

        self.header_fill = PatternFill(
            start_color=HEADER_FILL_COLOR,
            end_color=HEADER_FILL_COLOR,
            fill_type="solid",
        )

        self.header_font = Font(bold=True, size=12)
        self.center_align = Alignment(horizontal="center", vertical="center")

        # Границы
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Анализаторы различий
        self.analyzers = [
            ListDifferenceAnalyzer(
                separator=getattr(config.comparison, "list_separator", ",") if config else ","
            ),
            BasicDifferenceAnalyzer(),
        ]

    def format(self, result: ComparisonResult, output_path: Path, **options: Any) -> None:
        """Форматирует и сохраняет результат сравнения в Excel"""
        try:
            self.logger.info(f"Создание Excel отчета: {output_path}")

            rows = result.metadata.get("shape", (0, 0))[0]
            if rows > EXCEL_LARGE_ROW_THRESHOLD:
                self.logger.warning(
                    f"Большой объём данных ({rows} строк): формирование Excel-отчёта "
                    f"может быть медленным. Для больших таблиц рассмотрите HTML-отчёт "
                    f"(укажите расширение .html в имени выходного файла)."
                )

            # Создаем рабочую книгу
            workbook = Workbook()

            # Удаляем лист по умолчанию
            workbook.remove(workbook.active)

            # Добавляем листы с данными
            self._add_comparison_sheets(workbook, result, options)

            # Добавляем лист со сводкой
            self._add_summary_sheet(workbook, result)

            # Сохраняем файл
            workbook.save(output_path)

            self.logger.info(f"Excel отчет успешно создан: {output_path}")

        except Exception as e:
            self.logger.error(f"Ошибка создания Excel отчета: {e}", exc_info=True)
            raise ApplicationError(f"Не удалось создать отчет: {e}") from e

    def get_supported_formats(self) -> list[str]:
        """Возвращает список поддерживаемых форматов"""
        return self.SUPPORTED_FORMATS.copy()

    def _add_comparison_sheets(
        self, workbook: Workbook, result: ComparisonResult, options: dict[str, Any]
    ) -> None:
        """Добавляет листы с результатами сравнения"""
        file1_name = options.get("file1_name", "Файл 1")
        file2_name = options.get("file2_name", "Файл 2")

        sheet1_name = file1_name[:SHEET_NAME_MAX_LEN]
        sheet2_name = file2_name[:SHEET_NAME_MAX_LEN]

        # Создаем столбцы с различиями
        diff_column1 = self._create_differences_column(
            result.file1_data, result.file2_data, result.differences_mask
        )
        diff_column2 = self._create_differences_column(
            result.file2_data, result.file1_data, result.differences_mask
        )

        # Добавляем данные к DataFrame
        result1_df = result.file1_data.copy()
        result1_df["Различия"] = diff_column1

        result2_df = result.file2_data.copy()
        result2_df["Различия"] = diff_column2

        # Создаем листы
        self._create_data_sheet(workbook, sheet1_name, result1_df, result.differences_mask)
        self._create_data_sheet(workbook, sheet2_name, result2_df, result.differences_mask)

    def _create_data_sheet(
        self, workbook: Workbook, sheet_name: str, data_df: pd.DataFrame, mask: pd.DataFrame
    ) -> None:
        """Создает лист с данными и применяет форматирование"""
        worksheet = workbook.create_sheet(title=sheet_name)

        # Добавляем данные. openpyxl не умеет писать pd.NA (в отличие от
        # numpy.nan), а advanced-сравнение добивает короткую таблицу через
        # reindex(fill_value=pd.NA). Поэтому NA меняем на None — пустую ячейку.
        for row in dataframe_to_rows(data_df, index=False, header=True):
            worksheet.append([None if pd.isna(v) else v for v in row])

        # Применяем стили к заголовкам
        for cell in worksheet[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Применяем подсветку только к различающимся ячейкам.
        # Векторизация: numpy.nonzero возвращает позиции диффов одним проходом,
        # поэтому стоимость линейна от количества различий, а не от размера
        # таблицы. Это критично для больших листов с редкими изменениями.
        diff_rows, diff_cols = mask.values.nonzero()
        for row_pos, col_idx in zip(diff_rows, diff_cols, strict=True):
            cell = worksheet.cell(row=int(row_pos) + 2, column=int(col_idx) + 1)
            cell.fill = self.highlight_fill
            cell.border = self.thin_border

        # Автоподбор ширины столбцов
        self._auto_adjust_columns(worksheet)

    def _create_differences_column(
        self, base_df: pd.DataFrame, other_df: pd.DataFrame, mask: pd.DataFrame
    ) -> pd.Series:
        """Создает столбец с описанием различий.

        Iterates only over diff positions (via ``mask.values.nonzero``)
        instead of the full ``rows × cols`` grid. For sparse differences
        this is roughly O(diffs) vs the previous O(rows × cols).
        """
        cols = list(base_df.columns)
        base_values = base_df.values
        other_values = other_df.values
        diff_rows, diff_cols = mask.values.nonzero()

        by_row: dict[int, list[str]] = {}
        for r, c in zip(diff_rows, diff_cols, strict=True):
            col_name = cols[c]
            old_value = base_values[r, c]
            new_value = other_values[r, c]
            analyzer = self._find_analyzer(old_value, new_value)
            try:
                detail = analyzer.analyze(old_value, new_value, col_name)
                message = f"{col_name}: {detail.description}"
            except Exception as e:
                self.logger.warning(f"Ошибка анализа различий для {col_name}: {e}")
                message = f"{col_name}: {old_value} → {new_value}"
            by_row.setdefault(int(r), []).append(message)

        n_rows = len(base_df)
        descriptions = [" | ".join(by_row.get(i, [])) for i in range(n_rows)]
        return pd.Series(descriptions, index=base_df.index, name="Различия")

    def _find_analyzer(self, old_value: Any, new_value: Any) -> IDifferenceAnalyzer:
        """Находит подходящий анализатор для значений"""
        for analyzer in self.analyzers:
            if analyzer.can_analyze(old_value, new_value):
                return analyzer

        # Возвращаем базовый анализатор, если никто не подошел
        return self.analyzers[-1]

    def _add_summary_sheet(self, workbook: Workbook, result: ComparisonResult) -> None:
        """Добавляет лист со сводкой результатов"""
        worksheet = workbook.create_sheet(title="Сводка", index=0)

        # Заголовок
        worksheet["A1"] = "Сводка сравнения файлов"
        worksheet["A1"].font = Font(bold=True, size=16)
        worksheet["A1"].alignment = self.center_align
        worksheet.merge_cells("A1:B1")

        # Статистика
        row = 3
        stats = [
            ("Общее количество ячеек:", result.metadata["total_cells"]),
            ("Ячеек с различиями:", result.metadata["different_cells"]),
            ("Процент схожести:", f"{result.metadata['similarity_percentage']:.2f}%"),
            (
                "Размер данных:",
                f"{result.metadata['shape'][0]} строк × {result.metadata['shape'][1]} столбцов",
            ),
        ]

        for label, value in stats:
            worksheet[f"A{row}"] = label
            worksheet[f"B{row}"] = value
            worksheet[f"A{row}"].font = Font(bold=True)
            row += 1

        # Различия по столбцам
        row += 2
        worksheet[f"A{row}"] = "Различия по столбцам:"
        worksheet[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        # Заголовки таблицы
        worksheet[f"A{row}"] = "Столбец"
        worksheet[f"B{row}"] = "Количество различий"
        worksheet[f"C{row}"] = "Процент различий"

        for col in ["A", "B", "C"]:
            worksheet[f"{col}{row}"].font = self.header_font
            worksheet[f"{col}{row}"].fill = self.header_fill
            worksheet[f"{col}{row}"].border = self.thin_border

        row += 1

        # Статистика по столбцам
        for col_name in result.differences_mask.columns:
            diff_count = result.differences_mask[col_name].sum()
            diff_percentage = (diff_count / len(result.differences_mask)) * 100

            worksheet[f"A{row}"] = col_name
            worksheet[f"B{row}"] = diff_count
            worksheet[f"C{row}"] = f"{diff_percentage:.1f}%"

            for col in ["A", "B", "C"]:
                worksheet[f"{col}{row}"].border = self.thin_border

            row += 1

        # Автоподбор ширины столбцов
        self._auto_adjust_columns(worksheet)

    def _auto_adjust_columns(self, worksheet: Any) -> None:
        """Подбирает ширину столбцов по выборке первых строк.

        Раньше функция проходила по ``worksheet.columns``, материализуя КАЖДУЮ
        ячейку листа (для листа 58k×19 — около 1.1 млн ячеек на лист) и вызывая
        ``str()``+``len()`` для каждой; это и было основной стоимостью при
        формировании больших отчётов. Теперь сканируем только первые
        ``COLUMN_WIDTH_SAMPLE_ROWS`` строк (+ строка заголовков): ширина всё
        равно лишь визуальная оценка, а стоимость становится практически
        постоянной и не зависит от числа строк.
        """
        last_row = min(worksheet.max_row, COLUMN_WIDTH_SAMPLE_ROWS + 1)
        widths: dict[str, int] = {}

        for row in worksheet.iter_rows(min_row=1, max_row=last_row):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                value = cell.value
                if value is None:
                    continue
                length = len(str(value))
                letter = cell.column_letter
                if length > widths.get(letter, 0):
                    widths[letter] = length

        for letter, max_length in widths.items():
            worksheet.column_dimensions[letter].width = min(max_length + 2, 50)
