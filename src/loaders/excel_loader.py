"""
Загрузчик Excel файлов
"""

import logging
from pathlib import Path
from typing import Any, ClassVar, Literal, cast

import pandas as pd
from openpyxl import load_workbook

from ..core.exceptions import FileLoadError
from ..core.interfaces import IFileLoader

ExcelEngine = Literal["openpyxl", "xlrd"]


class ExcelFileLoader(IFileLoader):
    """Загрузчик для Excel файлов (.xlsx, .xls)"""

    SUPPORTED_EXTENSIONS: ClassVar[list[str]] = [".xlsx", ".xls"]

    def __init__(self) -> None:
        self.logger = logging.getLogger(self.__class__.__name__)

    def can_load(self, file_path: Path) -> bool:
        """Проверяет, может ли загрузчик обработать данный файл"""
        return file_path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def load(self, file_path: Path, **kwargs: Any) -> pd.DataFrame:
        """Загружает Excel файл и возвращает DataFrame.

        Supported kwargs:
            sheet_name: int | str — sheet to read (default 0)
            header: int — header row index (default 0)
            read_only: bool — bypass pandas for large xlsx files, reading
                via openpyxl streaming mode (default False). Significantly
                reduces memory usage for multi-MB files at a small CPU cost.
        """
        if not file_path.exists():
            raise FileLoadError(str(file_path), "Файл не существует")

        if not self.can_load(file_path):
            raise FileLoadError(
                str(file_path),
                f"Неподдерживаемое расширение. Поддерживаемые: {', '.join(self.SUPPORTED_EXTENSIONS)}",
            )

        try:
            sheet_name = kwargs.get("sheet_name", 0)
            header = kwargs.get("header", 0)
            read_only = bool(kwargs.get("read_only", False))

            self.logger.info(f"Загрузка файла: {file_path} (read_only={read_only})")

            if read_only and file_path.suffix.lower() == ".xlsx":
                df = self._load_xlsx_readonly(file_path, sheet_name, header)
            else:
                engine: ExcelEngine = "openpyxl" if file_path.suffix.lower() == ".xlsx" else "xlrd"
                df = cast(
                    pd.DataFrame,
                    pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        header=header,
                        engine=engine,
                    ),
                )

            self.logger.info(f"Успешно загружен файл {file_path.name}: {df.shape}")

            if df.empty:
                raise FileLoadError(str(file_path), "Файл пустой или не содержит данных")

            return df

        except PermissionError as e:
            raise FileLoadError(str(file_path), "Закройте файл Excel и попробуйте снова") from e
        except pd.errors.EmptyDataError as e:
            raise FileLoadError(str(file_path), "Файл не содержит данных") from e
        except Exception as e:
            self.logger.error(f"Ошибка загрузки файла {file_path}: {e}")
            raise FileLoadError(str(file_path), str(e)) from e

    @staticmethod
    def _load_xlsx_readonly(file_path: Path, sheet_name: int | str, header: int) -> pd.DataFrame:
        """Stream xlsx through openpyxl read_only mode to avoid materialising
        the full workbook in memory. Much cheaper for large files.
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            # Skip rows above the header row.
            for _ in range(header):
                next(rows_iter, None)
            header_row = next(rows_iter, None)
            if header_row is None:
                return pd.DataFrame()
            columns = [str(c) if c is not None else "" for c in header_row]
            data = list(rows_iter)
            return pd.DataFrame(data, columns=columns)
        finally:
            wb.close()

    def get_supported_extensions(self) -> list[str]:
        """Возвращает список поддерживаемых расширений"""
        return self.SUPPORTED_EXTENSIONS.copy()

    def get_sheet_names(self, file_path: Path) -> list[str]:
        """Получает список листов в Excel файле"""
        try:
            if file_path.suffix.lower() == ".xlsx":
                wb = load_workbook(file_path, read_only=True, data_only=True)
                try:
                    return list(wb.sheetnames)
                finally:
                    wb.close()
            # For legacy .xls go through pandas; context-manage the ExcelFile.
            with pd.ExcelFile(file_path, engine="xlrd") as xf:
                return [str(name) for name in xf.sheet_names]
        except Exception as e:
            self.logger.error(f"Ошибка получения списка листов из {file_path}: {e}")
            return []

    def preview_data(self, file_path: Path, max_rows: int = 5) -> pd.DataFrame:
        """Предварительный просмотр данных из файла"""
        try:
            engine: ExcelEngine = "openpyxl" if file_path.suffix.lower() == ".xlsx" else "xlrd"
            return pd.read_excel(file_path, nrows=max_rows, engine=engine)
        except Exception as e:
            self.logger.error(f"Ошибка предварительного просмотра {file_path}: {e}")
            return pd.DataFrame()
